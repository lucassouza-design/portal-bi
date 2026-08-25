"""
github_sharepoint_atualizar_dashboard.py
Fluxo:
  1. Autentica no Graph API (client credentials)
  2. Resolve o site e o drive certo pela biblioteca "Técnica - Documentos"
  3. Baixa Gerenciador_Atualizado_Ativo_2026.xlsx de PCM/Controle
  4. Lê "GESTÃO INPUTS" e "RESUMO SEMANAL" (parsing em streaming, read_only)
     -> valida o cabeçalho de cada coluna mapeada antes de ler os dados
  5. Atualiza dashboard_mcpack.html (RAW, RAW_SEMANAL, DATA_ATUALIZACAO, etc.)
  6. Envia e-mail de status (sucesso ou erro)
Env vars esperadas (já existem como Secrets no repo):
  AZURE_TENANT_ID, AZURE_CLIENT_ID, AZURE_CLIENT_SECRET
  EMAIL_REMETENTE, EMAIL_SENHA
"""
import io
import os
import re
import json
import smtplib
import unicodedata
import requests
from datetime import datetime, date, timezone, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import openpyxl
# ── Azure / Graph ────────────────────────────────────────────────────────────
AZ_TENANT_ID     = os.environ["AZURE_TENANT_ID"]
AZ_CLIENT_ID     = os.environ["AZURE_CLIENT_ID"]
AZ_CLIENT_SECRET = os.environ["AZURE_CLIENT_SECRET"]
SP_SITE_HOST    = os.environ["SP_SITE_HOST"]
SP_SITE_PATH    = os.environ["SP_SITE_PATH"]
SP_FILE_PATH    = os.environ["SP_FILE_PATH"]
# ── E-mail ────────────────────────────────────────────────────────────────────
EMAIL_REMETENTE     = os.environ["EMAIL_REMETENTE"]
EMAIL_SENHA          = os.environ["EMAIL_SENHA"]
EMAIL_DESTINATARIOS = [e.strip() for e in os.environ["EMAIL_DESTINATARIOS"].split(",") if e.strip()]
SMTP_HOST = "smtp.office365.com"
SMTP_PORT = 587
# ── Caminhos locais ──────────────────────────────────────────────────────────
HTML_PATH = "dashboard-atendimentos/atendimentos.html"  # caminho no repo portal-bi
# ── Fontes na planilha ────────────────────────────────────────────────────────
SHEET_INPUTS  = "GESTÃO INPUTS"
SHEET_SEMANAL = "RESUMO SEMANAL"
COL_MAP_INPUTS = {
    0: "id", 1: "data", 3: "cliente", 4: "unidade", 5: "cidade", 6: "pais",
    9: "origem", 11: "servico", 12: "tipo", 13: "modalidade", 14: "equip",
    15: "garantia", 16: "reinc", 17: "desc", 18: "status", 19: "tec1",
    20: "tec2", 21: "inicio", 22: "fim", 23: "diasTec", 24: "totalDias",
    35: "obs", 39: "idLink", 40: "dataAbert", 41: "dataEncerr",
    44: "dataEntrega", 45: "atraso", 46: "statusAuvo", 47: "dataEnvio",
}
DATE_FIELDS = {"data", "inicio", "fim", "dataAbert", "dataEncerr", "dataEntrega", "dataEnvio"}
INT_FIELDS  = {"id", "diasTec", "totalDias", "atraso", "reinc"}
FIELDS = [
    "id", "data", "cliente", "unidade", "cidade", "pais", "origem", "servico",
    "tipo", "modalidade", "equip", "garantia", "reinc", "desc", "status", "tec1", "tec2",
    "inicio", "fim", "diasTec", "totalDias", "obs", "idLink", "dataAbert",
    "dataEncerr", "dataEntrega", "atraso", "statusAuvo", "dataEnvio",
]
SEMANAL_COL_MAP = {
    0: "id", 1: "idPai", 2: "tec1", 3: "tec2", 4: "cliente", 5: "pais",
    6: "periodo", 7: "tipo", 8: "modalidade", 9: "status", 10: "relatorio",
    11: "envioCliente", 12: "obs",
}
# ── Rótulos esperados p/ validação de cabeçalho ──────────────────────────────
# GESTÃO INPUTS tem cabeçalho em 2 linhas (título principal + sub-título para
# colunas agrupadas, ex.: "AGENDA DE PLANEJAMENTO" -> "Início"/"Fim"). A
# validação aceita o rótulo em qualquer uma das duas linhas.
HEADER_LABELS_INPUTS = {
    "id": "ID REGISTRO",
    "data": "DATA REGISTRO",
    "cliente": "CLIENTE",
    "unidade": "UNIDADE",
    "cidade": "CIDADE",
    "pais": "PAIS",
    "origem": "ORIGEM",
    "servico": "SERVICO",
    "tipo": "TIPO DE SERVICO",
    "modalidade": "MODALIDADE",
    "equip": "EQUIPAMENTO",
    "garantia": "GARANTIA",
    "reinc": "REINCIDENCIA",
    "desc": "DESCRICAO",
    "status": "STATUS",
    "tec1": "TECNICO 1",
    "tec2": "TECNICO 2",
    "inicio": "INICIO",
    "fim": "FIM",
    "diasTec": "DIAS/TECNICO",
    "totalDias": "TOTAL DIAS",
    "obs": "OBSERVACOES",
    "idLink": "ID - LINK",
    "dataAbert": "DATA ABETURA",
    "dataEncerr": "DATA ENCERRAMENTO",
    "dataEntrega": "DATA ENTREGA",
    "atraso": "ATRASO",
    "statusAuvo": "STATUS",
    "dataEnvio": "DATA ENVIO",
}
HEADER_LABELS_SEMANAL = {
    "id": "ID REGISTRO",
    "idPai": "ID PAI",
    "tec1": "TECNICO 1",
    "tec2": "TECNICO 2",
    "cliente": "CLIENTE",
    "pais": "PAIS",
    "periodo": "PERIODO",
    "tipo": "TIPO DE SERVICO",
    "modalidade": "MODALIDADE",
    "status": "STATUS",
    "relatorio": "RELATORIO",
    "envioCliente": "ENVIO CLIENTE",
    "obs": "OBSERVACAO",
}
# ═════════════════════════════════════════════════════════════════════════════
# GRAPH — auth, resolve site/drive, download
# ═════════════════════════════════════════════════════════════════════════════
def get_graph_token() -> str:
    url = f"https://login.microsoftonline.com/{AZ_TENANT_ID}/oauth2/v2.0/token"
    resp = requests.post(
        url,
        data={
            "grant_type": "client_credentials",
            "client_id": AZ_CLIENT_ID,
            "client_secret": AZ_CLIENT_SECRET,
            "scope": "https://graph.microsoft.com/.default",
        },
        timeout=30,
    )
    resp.raise_for_status()
    token = resp.json().get("access_token")
    if not token:
        raise ValueError(f"Token Graph não retornado: {resp.text}")
    print("[GRAPH] Token obtido.")
    return token
def get_site_id(token: str) -> str:
    url = f"https://graph.microsoft.com/v1.0/sites/{SP_SITE_HOST}:{SP_SITE_PATH}"
    resp = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=30)
    resp.raise_for_status()
    site_id = resp.json().get("id")
    print(f"[SP] Site ID: {site_id}")
    return site_id
def get_drive_id(token: str, site_id: str) -> str:
    """'Documentos Compartilhados' é a biblioteca padrão do site — usa o drive padrão direto."""
    url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive"
    resp = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=30)
    resp.raise_for_status()
    drive = resp.json()
    print(f"[SP] Drive padrão: {drive.get('name')} ({drive['id']})")
    return drive["id"]
def download_sp_file(token: str, drive_id: str, file_path: str) -> bytes:
    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{file_path}:/content"
    resp = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=120)
    resp.raise_for_status()
    print(f"[SP] Download OK: {file_path} ({len(resp.content)} bytes)")
    return resp.content
# ═════════════════════════════════════════════════════════════════════════════
# VALIDAÇÃO DE CABEÇALHO
# ═════════════════════════════════════════════════════════════════════════════
def _norm(s):
    """Maiúsculas, sem acento, sem espaço/pontuação supérflua — pra comparar
    rótulos de forma tolerante a pequenas variações de formatação."""
    if s is None:
        return ""
    s = str(s).strip().upper()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^A-Z0-9/]+", " ", s).strip()
    return s
def validate_headers(rows, col_map, header_labels, sheet_name):
    """
    rows: lista de tuplas (linhas de cabeçalho candidatas, ex.: [linha_titulo,
          linha_subtitulo]) — o rótulo esperado pode aparecer em qualquer uma.
    Lança RuntimeError listando TODAS as colunas divergentes de uma vez, pra
    não descobrir os problemas um de cada vez a cada re-execução.
    """
    erros = []
    for idx, field in col_map.items():
        esperado = header_labels.get(field)
        if esperado is None:
            continue
        esperado_norm = _norm(esperado)
        encontrados = []
        ok = False
        for row in rows:
            valor = row[idx] if idx < len(row) else None
            encontrados.append(valor)
            if esperado_norm and esperado_norm in _norm(valor):
                ok = True
                break
        if not ok:
            erros.append(
                f"  coluna {idx} (campo '{field}'): esperava conter \"{esperado}\", "
                f"achei {encontrados!r}"
            )
    if erros:
        raise RuntimeError(
            f"Cabeçalho da aba '{sheet_name}' não bate com o mapeamento esperado "
            f"(planilha pode ter mudado de estrutura):\n" + "\n".join(erros)
        )
# ═════════════════════════════════════════════════════════════════════════════
# PARSING (streaming — read_only evita carregar estilos/formatação)
# ═════════════════════════════════════════════════════════════════════════════
def to_date_str(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    return s if s != "" else None
def to_int_or_none(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        try:
            return int(v)
        except (ValueError, OverflowError):
            return v
    s = str(v).strip()
    try:
        return int(float(s))
    except ValueError:
        return s
def clean_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s != "" else None
def find_header_rows(ws, sheet_name):
    """
    Retorna (data_start_row, linha_titulo, linha_subtitulo) para a aba
    GESTÃO INPUTS: a linha de título é a que começa com "ID REGISTRO", a
    linha seguinte é o sub-título (colunas agrupadas), e os dados começam
    2 linhas depois do título.
    """
    linha_titulo = None
    linha_titulo_idx = None
    linha_subtitulo = None
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        v = row[0] if row else None
        if v is not None and str(v).strip().upper().startswith("ID REGISTRO"):
            linha_titulo = row
            linha_titulo_idx = row_idx
            break
    if linha_titulo is None:
        raise RuntimeError(
            f"Não encontrei a linha de cabeçalho 'ID REGISTRO' na aba '{sheet_name}'. "
            "A estrutura da planilha pode ter mudado — verifique manualmente."
        )
    # linha de sub-título logo abaixo (pode não existir/estar vazia — tudo bem)
    for row_idx, row in enumerate(ws.iter_rows(min_row=linha_titulo_idx + 1,
                                                max_row=linha_titulo_idx + 1,
                                                values_only=True)):
        linha_subtitulo = row
    data_start_row = linha_titulo_idx + 2
    return data_start_row, linha_titulo, linha_subtitulo or ()
def load_gestao_inputs(wb):
    ws = wb[SHEET_INPUTS]
    data_start_row, linha_titulo, linha_subtitulo = find_header_rows(ws, SHEET_INPUTS)
    validate_headers([linha_titulo, linha_subtitulo], COL_MAP_INPUTS, HEADER_LABELS_INPUTS, SHEET_INPUTS)
    records = []
    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        if row is None or row[0] is None:
            continue
        rec = {f: None for f in FIELDS}
        for idx, field in COL_MAP_INPUTS.items():
            if idx >= len(row):
                continue
            v = row[idx]
            if field in DATE_FIELDS:
                rec[field] = to_date_str(v)
            elif field in INT_FIELDS:
                rec[field] = to_int_or_none(v)
            else:
                rec[field] = clean_str(v)
        records.append(rec)
    return records
def load_resumo_semanal(wb):
    ws = wb[SHEET_SEMANAL]
    semana = None
    ano = None
    header_row = None
    header_row_idx = None
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        v = row[0] if row else None
        if v is not None and str(v).strip().upper().startswith("SEMANA"):
            semana = row[2] if len(row) > 2 else None
            ano = row[4] if len(row) > 4 else None
        if v is not None and str(v).strip().upper().startswith("ID REGISTRO"):
            header_row = row
            header_row_idx = row_idx
            break
    if header_row is None:
        raise RuntimeError(
            f"Não encontrei a linha de cabeçalho 'ID REGISTRO' na aba '{SHEET_SEMANAL}'. "
            "A estrutura da planilha pode ter mudado — verifique manualmente."
        )
    validate_headers([header_row], SEMANAL_COL_MAP, HEADER_LABELS_SEMANAL, SHEET_SEMANAL)
    data_start_row = header_row_idx + 1
    records = []
    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        if row is None or row[0] is None:
            continue
        rec = {}
        for idx, field in SEMANAL_COL_MAP.items():
            v = row[idx] if idx < len(row) else None
            if field in ("id", "idPai"):
                rec[field] = to_int_or_none(v)
            else:
                rec[field] = clean_str(v) if not (v == 0) else None
        records.append(rec)
    return semana, ano, records
def build_js_const(name, data):
    return f"const {name} = " + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";"
def update_html(html_path, records, semana, ano, semanal_records, output_path):
    with open(html_path, "r", encoding="utf-8") as f:
        html = f.read()
    new_raw = build_js_const("RAW", records)
    pattern = re.compile(r"const RAW\s*=\s*\[.*?\];", re.DOTALL)
    if not pattern.search(html):
        raise RuntimeError("Não encontrei 'const RAW = [...]' no HTML — verifique o arquivo de entrada.")
    html = pattern.sub(lambda m: new_raw, html, count=1)
    new_raw_semanal = build_js_const("RAW_SEMANAL", semanal_records)
    pattern_sem = re.compile(r"const RAW_SEMANAL\s*=\s*\[.*?\];", re.DOTALL)
    if pattern_sem.search(html):
        html = pattern_sem.sub(lambda m: new_raw_semanal, html, count=1)
    else:
        html = html.replace(new_raw, new_raw + "\n" + new_raw_semanal, 1)
    data_atualizacao = datetime.now(tz=timezone(timedelta(hours=-3))).strftime("%d/%m/%Y %H:%M")
    data_js = f'const DATA_ATUALIZACAO = "{data_atualizacao}";'
    pattern_data = re.compile(r'const DATA_ATUALIZACAO\s*=.*?;')
    if pattern_data.search(html):
        html = pattern_data.sub(lambda m: data_js, html, count=1)
    else:
        html = html.replace(new_raw_semanal, new_raw_semanal + "\n" + data_js, 1)
    semana_js = f"const SEMANA_ATUAL = {json.dumps(semana)};\nconst ANO_ATUAL = {json.dumps(ano)};"
    pattern_sw = re.compile(r"const SEMANA_ATUAL\s*=.*?;\s*\nconst ANO_ATUAL\s*=.*?;", re.DOTALL)
    if pattern_sw.search(html):
        html = pattern_sw.sub(lambda m: semana_js, html, count=1)
    else:
        html = html.replace(new_raw_semanal, new_raw_semanal + "\n" + semana_js, 1)
    total = len(records)
    html = re.sub(
        r"VISÃO GERAL — TODA A BASE \(\d+ REGISTROS\)",
        f"VISÃO GERAL — TODA A BASE ({total} REGISTROS)",
        html,
    )
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
# ═════════════════════════════════════════════════════════════════════════════
# E-MAIL
# ═════════════════════════════════════════════════════════════════════════════
def saudacao() -> str:
    hora = datetime.now(tz=timezone(timedelta(hours=-3))).hour
    if hora < 12:
        return "Bom dia"
    elif hora < 18:
        return "Boa tarde"
    return "Boa noite"
def enviar_email(sucesso: bool, total_registros: int = 0, erro: str = "") -> None:
    agora = datetime.now(tz=timezone(timedelta(hours=-3))).strftime("%d/%m/%Y %H:%M")
    if sucesso:
        assunto = f"✅ Gerenciador Ativos atualizado — {agora}"
        corpo = f"""
<p>{saudacao()} time,</p>
<p>O Gerenciador Ativos no portal BI foi atualizado em <b>{agora}</b>. Em alguns instantes os dados atualizados já estarão disponíveis.</p>
<p>Total de registros: <b>{total_registros}</b></p>
<p>Atenciosamente,<br>
Lucas Souza<br>
Analista de Dados — MCPack</p>
<p style="color:#888; font-size:12px; margin-top:16px;">
Este email foi enviado automaticamente por um sistema interno para uma lista de destinatários,
usando esta caixa de entrada apenas como remetente técnico. Não é necessário responder.
</p>
"""
    else:
        assunto = f"❌ Falha ao atualizar Gerenciador Ativos — {agora}"
        corpo = f"""
<p>{saudacao()} time,</p>
<p>Ocorreu um erro ao atualizar o Gerenciador Ativos no portal BI em <b>{agora}</b>.</p>
<p><b>Erro:</b> {erro}</p>
<p>Atenciosamente,<br>
Lucas Souza<br>
Analista de Dados — MCPack</p>
<p style="color:#888; font-size:12px; margin-top:16px;">
Este email foi enviado automaticamente por um sistema interno para uma lista de destinatários,
usando esta caixa de entrada apenas como remetente técnico. Não é necessário responder.
</p>
"""
    msg = MIMEMultipart("mixed")
    msg["Subject"] = assunto
    msg["From"] = EMAIL_REMETENTE
    msg["To"] = ", ".join(EMAIL_DESTINATARIOS)
    msg.attach(MIMEText(corpo, "html"))
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
        server.ehlo()
        server.starttls()
        server.login(EMAIL_REMETENTE, EMAIL_SENHA)
        server.sendmail(EMAIL_REMETENTE, EMAIL_DESTINATARIOS, msg.as_string())
    print(f"[EMAIL] Enviado para: {', '.join(EMAIL_DESTINATARIOS)}")
# ═════════════════════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════════════════════
def main():
    try:
        token = get_graph_token()
        site_id = get_site_id(token)
        drive_id = get_drive_id(token, site_id)
        xlsx_bytes = download_sp_file(token, drive_id, SP_FILE_PATH)
        print("Lendo planilha...")
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
        records = load_gestao_inputs(wb)
        print(f"[GESTÃO INPUTS] Registros lidos: {len(records)}")
        semana, ano, semanal_records = load_resumo_semanal(wb)
        print(f"[RESUMO SEMANAL] Semana {semana}/{ano} — {len(semanal_records)} registro(s)")
        update_html(HTML_PATH, records, semana, ano, semanal_records, HTML_PATH)
        print(f"Dashboard atualizado: {HTML_PATH}")
        enviar_email(True, total_registros=len(records))
    except Exception as e:
        print(f"[ERRO FATAL] {e}")
        try:
            enviar_email(False, erro=str(e))
        except Exception as email_err:
            print(f"[ERRO AO ENVIAR E-MAIL DE FALHA] {email_err}")
        raise
if __name__ == "__main__":
    main()
